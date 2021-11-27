from openpyxl import load_workbook

class Readxls():
   def read_data(self):
       wb = load_workbook("../case/case1.xlsx")
       # 获取所有表格(worksheet)的名字
       sheets = wb.sheetnames
       # 第一个表格的名称
       sheet_first = sheets[0]
       # 获取特定的worksheet
       ws = wb[sheet_first]

       # 获取表格所有行和列，两者都是可迭代的
       rows = ws.rows
       content = []
       # 迭代所有的行
       for row in rows:
           line = [col.value for col in row]
           content.append(line)
       print(content)
       return content
